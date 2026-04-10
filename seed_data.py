"""
Generate realistic mock data for the Snowball Developments reporting demo.

Creates three Excel files that mirror Juniper Square CSV export structure:
  - investor_roster.xlsx  (20 investors across 3 funds)
  - fund_performance.xlsx (3 tabs: GP1, GP2, CAD_FEEDER)
  - quarter_inputs.xlsx   (reporting period, FX, withholding, highlights)

Asset names are pulled from Snowball's known portfolio where possible.
All figures are plausible-but-fictional.
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook


# ---------- INVESTOR ROSTER ----------

INVESTORS = [
    # GP Fund #1 — 10 investors, fully called
    ("INV-001", "Harold Weinstein Family Trust", "GP1", "HNW", "US",
     1_000_000, 1_000_000, 180_000, 8.33,
     "hweinstein@example.com", "445 Park Ave\nNew York, NY 10022"),
    ("INV-002", "Thornton Capital Partners LLC", "GP1", "Family Office", "US",
     2_500_000, 2_500_000, 450_000, 20.83,
     "ir@thorntoncapital.com", "1 Rockefeller Plaza\nNew York, NY 10020"),
    ("INV-003", "Maria Delgado", "GP1", "HNW", "US",
     750_000, 750_000, 135_000, 6.25,
     "maria.delgado@example.com", "27 West 72nd St Apt 4B\nNew York, NY 10023"),
    ("INV-004", "Emerson Pension Trust", "GP1", "Institutional", "US",
     2_000_000, 2_000_000, 360_000, 16.67,
     "trustees@emersonpension.org", "100 Federal St\nBoston, MA 02110"),
    ("INV-005", "James O'Connor", "GP1", "HNW", "US",
     500_000, 500_000, 90_000, 4.17,
     "joconnor@example.com", "14 Elm Street\nGreenwich, CT 06830"),
    ("INV-006", "Blackstone Family Office LLC", "GP1", "Family Office", "US",
     1_500_000, 1_500_000, 270_000, 12.50,
     "office@blackstonefo.com", "280 Park Ave\nNew York, NY 10017"),
    ("INV-007", "Priya Venkatesan", "GP1", "HNW", "US",
     600_000, 600_000, 108_000, 5.00,
     "pvenkatesan@example.com", "220 Central Park South\nNew York, NY 10019"),
    ("INV-008", "Ridgeway Holdings LP", "GP1", "Family Office", "US",
     1_250_000, 1_250_000, 225_000, 10.42,
     "investments@ridgeway.com", "1114 Avenue of the Americas\nNew York, NY 10036"),
    ("INV-009", "Samuel Chen", "GP1", "HNW", "US",
     850_000, 850_000, 153_000, 7.08,
     "schen@example.com", "88 Morningside Dr\nNew York, NY 10027"),
    ("INV-010", "West Harbor Endowment", "GP1", "Institutional", "US",
     1_050_000, 1_050_000, 189_000, 8.75,
     "treasury@westharboredu.org", "55 Water Street\nNew York, NY 10041"),

    # GP Fund #2 — 5 investors, mid-raise (partial calls)
    ("INV-011", "Lighthouse Capital Partners", "GP2", "Family Office", "US",
     3_000_000, 1_500_000, 0, 30.00,
     "investments@lighthousecap.com", "600 Madison Ave\nNew York, NY 10022"),
    ("INV-012", "Robert Ashford", "GP2", "HNW", "US",
     1_000_000, 500_000, 0, 10.00,
     "rashford@example.com", "18 Gramercy Park South\nNew York, NY 10003"),
    ("INV-013", "Meridian Foundation", "GP2", "Institutional", "US",
     2_500_000, 1_250_000, 0, 25.00,
     "finance@meridianfound.org", "1290 Sixth Ave\nNew York, NY 10104"),
    ("INV-014", "Katherine Liu", "GP2", "HNW", "US",
     1_500_000, 750_000, 0, 15.00,
     "kliu@example.com", "1 Central Park West\nNew York, NY 10023"),
    ("INV-015", "Summit Ridge Capital", "GP2", "Family Office", "US",
     2_000_000, 1_000_000, 0, 20.00,
     "ir@summitridge.com", "9 West 57th St\nNew York, NY 10019"),

    # Canadian Feeder Fund — 5 investors
    ("INV-016", "Pierre Lavoie Family Trust", "CAD_FEEDER", "HNW", "Canada",
     1_200_000, 1_200_000, 216_000, 24.00,
     "plavoie@example.ca", "1 Place Ville Marie\nMontreal, QC H3B 2C4"),
    ("INV-017", "Maple Crest Holdings Inc.", "CAD_FEEDER", "Family Office", "Canada",
     1_500_000, 1_500_000, 270_000, 30.00,
     "admin@maplecrest.ca", "161 Bay Street\nToronto, ON M5J 2S1"),
    ("INV-018", "Jean-Luc Tremblay", "CAD_FEEDER", "HNW", "Canada",
     800_000, 800_000, 144_000, 16.00,
     "jltremblay@example.ca", "1250 René-Lévesque Blvd W\nMontreal, QC H3B 4W8"),
    ("INV-019", "BC Pacific Investment Trust", "CAD_FEEDER", "Institutional", "Canada",
     1_000_000, 1_000_000, 180_000, 20.00,
     "trust@bcpacific.ca", "666 Burrard Street\nVancouver, BC V6C 3P6"),
    ("INV-020", "Sarah Mackenzie", "CAD_FEEDER", "HNW", "Canada",
     500_000, 500_000, 90_000, 10.00,
     "smackenzie@example.ca", "181 Bay Street\nToronto, ON M5J 2T3"),
]


def build_investor_roster(path: Path):
    df = pd.DataFrame(INVESTORS, columns=[
        "Investor ID", "Investor Name", "Fund Vehicle", "Investor Type",
        "Country", "Commitment", "Total Contributed", "Total Distributed",
        "Ownership %", "Email", "Mailing Address",
    ])
    df.to_excel(path, index=False)


# ---------- FUND PERFORMANCE ----------

# GP Fund #1 — operating portfolio, 14 assets
GP1_ASSETS = [
    ("Eastern Park",          "CT", 180_000, 95, 1_620_000,  8.5, 21_000_000),
    ("Mountainside",          "NJ",  95_000, 98,   935_000,  6.2, 14_500_000),
    ("South Windsor",         "CT", 145_000, 97,  1_180_000, 11.3, 16_200_000),
    ("Danbury",               "CT", 125_000, 92,    870_000,  5.8, 11_800_000),
    ("Little Ferry",          "NJ",  68_000, 100,   720_000,  9.4, 10_500_000),
    ("Bridgeport",            "CT", 210_000, 88,  1_380_000,  4.2, 17_500_000),
    ("Stratford",             "CT", 155_000, 94,  1_050_000,  7.1, 13_200_000),
    ("Waterbury",             "CT", 135_000, 100,   985_000, 12.6, 12_100_000),
    ("Windsor Locks",         "CT", 175_000, 90,  1_220_000,  6.8, 14_700_000),
    ("East Hartford",         "CT", 115_000, 95,    810_000,  9.2, 10_300_000),
    ("Elizabeth",             "NJ",  85_000, 100,   925_000,  8.1, 12_800_000),
    ("Kearny",                "NJ",  72_000, 96,    780_000,  7.5, 10_900_000),
    ("Meriden",               "CT", 160_000, 82,  1_020_000,  3.4, 13_500_000),
    ("New Britain",           "CT",  80_000, 100,   615_000, 10.8,  8_400_000),
]

# GP Fund #2 — early stage, 3 assets acquired so far
GP2_ASSETS = [
    ("Norwalk Industrial",    "CT", 120_000, 100,   870_000,  0.0, 11_800_000),
    ("Paterson Logistics",    "NJ",  95_000,  85,   620_000,  0.0,  8_900_000),
    ("North Haven",           "CT", 105_000,  95,   745_000,  0.0,  9_700_000),
]

# Canadian Feeder tracks a subset of GP1's NAV (feeder structure)
CAD_FEEDER_ASSETS = GP1_ASSETS  # Feeder sees through to the GP fund portfolio


def build_fund_performance(path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    funds = [
        ("GP_Fund_1", {
            "Quarter-End NAV": 168_500_000,
            "Quarterly Distribution per Unit": 0.45,
            "YTD Net IRR": 14.2,
            "ITD Net IRR": 17.8,
            "YTD Equity Multiple": 1.08,
            "ITD Equity Multiple": 1.42,
            "Fund Unfunded Commitment": 0,
        }, GP1_ASSETS),
        ("GP_Fund_2", {
            "Quarter-End NAV": 10_250_000,
            "Quarterly Distribution per Unit": 0.00,
            "YTD Net IRR": 0.0,
            "ITD Net IRR": 0.0,
            "YTD Equity Multiple": 1.00,
            "ITD Equity Multiple": 1.00,
            "Fund Unfunded Commitment": 15_000_000,
        }, GP2_ASSETS),
        ("CAD_Feeder", {
            "Quarter-End NAV": 5_200_000,
            "Quarterly Distribution per Unit": 0.42,
            "YTD Net IRR": 13.5,
            "ITD Net IRR": 16.9,
            "YTD Equity Multiple": 1.07,
            "ITD Equity Multiple": 1.38,
            "Fund Unfunded Commitment": 0,
        }, CAD_FEEDER_ASSETS),
    ]

    for tab_name, summary, assets in funds:
        ws = wb.create_sheet(title=tab_name)
        # Summary block (7 rows, key-value)
        for i, (k, v) in enumerate(summary.items(), start=1):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)

        # Asset table header at row 10
        header = ["Asset Name", "State", "SF", "Occupancy %",
                  "In-Place NOI", "YoY NOI Change %", "Valuation Mark"]
        for col, h in enumerate(header, start=1):
            ws.cell(row=10, column=col, value=h)

        for row_idx, asset in enumerate(assets, start=11):
            for col_idx, val in enumerate(asset, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(path)


# ---------- QUARTER INPUTS ----------

def build_quarter_inputs(path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    config = wb.create_sheet(title="Config")
    config_rows = [
        ("Reporting Period", "Q1 2026"),
        ("FX Rate USD/CAD", 1.3650),
        ("CAD Withholding Rate", 0.15),
    ]
    for row_idx, (k, v) in enumerate(config_rows, start=1):
        config.cell(row=row_idx, column=1, value=k)
        config.cell(row=row_idx, column=2, value=v)

    highlights = wb.create_sheet(title="Highlights")
    highlights.cell(row=1, column=1, value="GP1")
    highlights.cell(row=1, column=2, value="GP2")
    highlights.cell(row=1, column=3, value="CAD_FEEDER")

    gp1_h = [
        "Portfolio occupancy reached 94% on a trailing basis, with ~350K SF of leases pending execution that will push occupancy toward 98% by Q2.",
        "Signed a new 10-year lease at South Windsor at $8.75 PSF NNN, a 42% mark-to-market premium over the prior in-place rent.",
        "Solar program expanded to 9 assets totaling 500K+ SF, now generating $245K in annual income at ~$0.50 PSF — none of which was modeled at acquisition.",
        "Completed the South Windsor refinancing at $9.5M against a $9.0M total cost basis, returning equity to the partnership.",
        "Waterbury and Meriden executed 4 acres of outdoor storage leases (IOS), adding incremental cash flow with minimal capex.",
    ]
    gp2_h = [
        "Three initial assets acquired in Q1: Norwalk Industrial, Paterson Logistics, and North Haven, totaling 320K SF for approximately $30.4M.",
        "Pipeline includes 6 assets under contract representing 600K SF, with closings anticipated between February and April.",
        "Fund is on track to be fully deployed by the end of 2026, supporting launch planning for GP Fund #2 follow-on vehicles.",
        "Value-add thesis unchanged: acquire land-heavy assets at below-replacement-cost basis in the Tri-State industrial market.",
    ]
    cad_h = [
        "The Canadian Feeder participated pro rata in GP Fund #1's Q1 activity, including the South Windsor refinancing and solar program expansion.",
        "FX conditions remained stable through the quarter at approximately 1.365 USD/CAD.",
        "No changes to the Canadian non-resident withholding rate of 15% on distributions.",
        "Eastern Canada expansion under evaluation for potential future feeder vehicles.",
    ]

    for i, h in enumerate(gp1_h, start=2):
        highlights.cell(row=i, column=1, value=h)
    for i, h in enumerate(gp2_h, start=2):
        highlights.cell(row=i, column=2, value=h)
    for i, h in enumerate(cad_h, start=2):
        highlights.cell(row=i, column=3, value=h)

    wb.save(path)


def build_all(data_dir: Path):
    data_dir.mkdir(parents=True, exist_ok=True)
    build_investor_roster(data_dir / "investor_roster.xlsx")
    build_fund_performance(data_dir / "fund_performance.xlsx")
    build_quarter_inputs(data_dir / "quarter_inputs.xlsx")


if __name__ == "__main__":
    import sys
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data")
    build_all(out)
    print(f"Mock data written to {out}/")
